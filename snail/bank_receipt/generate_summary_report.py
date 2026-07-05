#!/usr/bin/env python3
"""
生成银行水单统计分析报告（千元/万元单位）
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import PieChart, BarChart, Reference
from datetime import datetime

def generate_summary_report():
    excel_path = r'D:\code\code\银行水单信息汇总.xlsx'
    
    # 加载原始数据
    wb = openpyxl.load_workbook(excel_path)
    ws_source = wb.active
    
    # 创建新的工作簿用于统计报告
    wb_report = openpyxl.Workbook()
    
    # ========== Sheet 1: 按付款公司统计（万元） ==========
    ws_company = wb_report.active
    ws_company.title = "按付款公司统计"
    
    # 统计公司数据
    company_stats = {}
    for row in ws_source.iter_rows(min_row=2, values_only=True):
        company = row[4]  # 公司名称
        amount = row[3]   # 金额
        
        if company and amount and isinstance(amount, (int, float)):
            if company not in company_stats:
                company_stats[company] = {'count': 0, 'amount': 0}
            company_stats[company]['count'] += 1
            company_stats[company]['amount'] += amount
    
    # 写入表头
    headers = ['付款公司', '笔数', '总金额（元）', '总金额（万元）', '总金额（千元）', '平均金额（元）']
    ws_company.append(headers)
    for cell in ws_company[1]:
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    # 写入数据（按金额降序排序）
    sorted_companies = sorted(company_stats.items(), key=lambda x: x[1]['amount'], reverse=True)
    for company, stats in sorted_companies:
        amount_wan = stats['amount'] / 10000  # 万元
        amount_qian = stats['amount'] / 1000  # 千元
        avg_amount = stats['amount'] / stats['count'] if stats['count'] > 0 else 0
        
        ws_company.append([
            company,
            stats['count'],
            stats['amount'],
            round(amount_wan, 2),
            round(amount_qian, 2),
            round(avg_amount, 2)
        ])
    
    # 添加总计行
    total_row = ['总计', '', '', '', '', '']
    ws_company.append(total_row)
    last_row = ws_company.max_row
    ws_company.cell(last_row, 1).value = '总计'
    ws_company.cell(last_row, 2).value = sum(s['count'] for s in company_stats.values())
    ws_company.cell(last_row, 3).value = sum(s['amount'] for s in company_stats.values())
    ws_company.cell(last_row, 4).value = round(sum(s['amount'] for s in company_stats.values()) / 10000, 2)
    ws_company.cell(last_row, 5).value = round(sum(s['amount'] for s in company_stats.values()) / 1000, 2)
    
    for cell in ws_company[last_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    # 设置列宽
    ws_company.column_dimensions['A'].width = 35
    ws_company.column_dimensions['B'].width = 10
    ws_company.column_dimensions['C'].width = 18
    ws_company.column_dimensions['D'].width = 18
    ws_company.column_dimensions['E'].width = 18
    ws_company.column_dimensions['F'].width = 18
    
    # ========== Sheet 2: 按月份统计 ==========
    ws_month = wb_report.create_sheet("按月份统计")
    
    month_stats = {}
    for row in ws_source.iter_rows(min_row=2, values_only=True):
        date_str = row[2]  # 日期
        amount = row[3]    # 金额
        
        if date_str and amount and isinstance(amount, (int, float)):
            # 提取月份
            try:
                if isinstance(date_str, str) and len(date_str) >= 7:
                    month = date_str[:7]  # YYYY-MM格式
                else:
                    continue
                
                if month not in month_stats:
                    month_stats[month] = {'count': 0, 'amount': 0}
                month_stats[month]['count'] += 1
                month_stats[month]['amount'] += amount
            except:
                pass
    
    # 写入月份统计
    ws_month.append(['月份', '笔数', '总金额（元）', '总金额（万元）', '总金额（千元）'])
    for cell in ws_month[1]:
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    for month in sorted(month_stats.keys()):
        stats = month_stats[month]
        ws_month.append([
            month,
            stats['count'],
            stats['amount'],
            round(stats['amount'] / 10000, 2),
            round(stats['amount'] / 1000, 2)
        ])
    
    # 添加总计行
    ws_month.append(['总计', '', '', '', ''])
    last_row = ws_month.max_row
    ws_month.cell(last_row, 1).value = '总计'
    ws_month.cell(last_row, 2).value = sum(s['count'] for s in month_stats.values())
    ws_month.cell(last_row, 3).value = sum(s['amount'] for s in month_stats.values())
    ws_month.cell(last_row, 4).value = round(sum(s['amount'] for s in month_stats.values()) / 10000, 2)
    ws_month.cell(last_row, 5).value = round(sum(s['amount'] for s in month_stats.values()) / 1000, 2)
    
    for cell in ws_month[last_row]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
    
    ws_month.column_dimensions['A'].width = 12
    ws_month.column_dimensions['B'].width = 10
    ws_month.column_dimensions['C'].width = 18
    ws_month.column_dimensions['D'].width = 18
    ws_month.column_dimensions['E'].width = 18
    
    # ========== Sheet 3: 按项目类型统计 ==========
    ws_project = wb_report.create_sheet("按项目类型统计")
    
    project_stats = {}
    for row in ws_source.iter_rows(min_row=2, values_only=True):
        project = row[5]  # 项目名称
        amount = row[3]   # 金额
        
        if project and amount and isinstance(amount, (int, float)):
            if project not in project_stats:
                project_stats[project] = {'count': 0, 'amount': 0}
            project_stats[project]['count'] += 1
            project_stats[project]['amount'] += amount
    
    # 写入项目统计
    ws_project.append(['项目名称', '笔数', '总金额（元）', '总金额（万元）'])
    for cell in ws_project[1]:
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    
    sorted_projects = sorted(project_stats.items(), key=lambda x: x[1]['amount'], reverse=True)
    for project, stats in sorted_projects:
        ws_project.append([
            project,
            stats['count'],
            stats['amount'],
            round(stats['amount'] / 10000, 2)
        ])
    
    ws_project.column_dimensions['A'].width = 40
    ws_project.column_dimensions['B'].width = 10
    ws_project.column_dimensions['C'].width = 18
    ws_project.column_dimensions['D'].width = 18
    
    # 保存报告
    report_path = r'D:\code\code\银行水单统计分析报告.xlsx'
    wb_report.save(report_path)
    
    print("=" * 70)
    print("[成功] 统计分析报告已生成！")
    print("=" * 70)
    print(f"\n报告文件：{report_path}")
    print(f"\n包含以下工作表：")
    print(f"  1. 按付款公司统计（含万元、千元单位）")
    print(f"  2. 按月份统计")
    print(f"  3. 按项目类型统计")
    
    # 显示前10名公司（万元）
    print(f"\n前10名付款公司（按金额排序）：")
    print(f"{'='*70}")
    print(f"{'排名':<6}{'公司名称':<35}{'笔数':<8}{'金额(万元)':<15}{'金额(千元)':<15}")
    print(f"{'='*70}")
    for idx, (company, stats) in enumerate(sorted_companies[:10], 1):
        amount_wan = stats['amount'] / 10000
        amount_qian = stats['amount'] / 1000
        print(f"{idx:<6}{company:<35}{stats['count']:<8}{amount_wan:<15.2f}{amount_qian:<15.2f}")

if __name__ == '__main__':
    generate_summary_report()