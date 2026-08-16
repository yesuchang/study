# -*- coding: utf-8 -*-
"""
company_doc.py - 解析企业信用报告 Word 文档 (.docx)

功能：
1. 用 pandoc (pypandoc) 将 docx 转换为 HTML
2. 解析单个 docx 文件，或目录下全部 docx（企业信用报告）
3. 从每份报告提取：
   - 公司名称、法定代表人（工商信息表）
   - 股东信息（发起人名称、持股比例、认缴出资额、认缴出资日期、首次持股日期）
   - 主要人员（姓名、职务、持股比例）
4. 以人名为索引合并所有公司信息，输出 xlsx（1 个 sheet）
   同一人出现在多家公司时，各列以换行分隔对应各公司信息

依赖：
  pip install pypandoc pypandoc_binary openpyxl
"""

import argparse
import json
import logging
import os
import re
import sys
from collections import OrderedDict
from html.parser import HTMLParser
from pathlib import Path

try:
    import pypandoc
except ImportError:
    sys.exit("缺少依赖 pypandoc，请先安装: pip install pypandoc pypandoc_binary")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("缺少依赖 openpyxl，请先安装: pip install openpyxl")


# ---------- 日志 ----------

logger = logging.getLogger("company_doc")
logger.addHandler(logging.NullHandler())


def setup_logging(log_file: str, verbose: bool = False):
    """配置日志：同时输出到文件和控制台"""
    logger.setLevel(logging.DEBUG)

    # 文件日志（UTF-8 编码，避免中文乱码）
    file_handler = logging.FileHandler(log_file, encoding="utf-8")
    file_handler.setLevel(logging.DEBUG)
    file_fmt = logging.Formatter(
        "%(asctime)s [%(levelname)s] %(message)s", datefmt="%Y-%m-%d %H:%M:%S"
    )
    file_handler.setFormatter(file_fmt)
    logger.addHandler(file_handler)

    # 控制台日志
    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO if verbose else logging.WARNING)
    console_fmt = logging.Formatter("[%(levelname)s] %(message)s")
    console_handler.setFormatter(console_fmt)
    logger.addHandler(console_handler)


# ---------- HTML 解析 ----------

class ReportHTMLParser(HTMLParser):
    """解析 pandoc 生成的 HTML，按章节收集表格数据"""

    def __init__(self):
        super().__init__()
        self.sections = OrderedDict()   # 章节名 -> [table, ...]，table = [row, ...]，row = [cell, ...]
        self.current_section = None
        self._h_tag = None             # 当前所在 h1/h2/h3
        self._h_text = []              # 标题文本累积

        self.in_table = False
        self.in_row = False
        self.in_cell = False
        self.current_row = []
        self.current_cell_text = []
        self.current_table = []

    # ---- 标签开始 ----
    def handle_starttag(self, tag, attrs):
        if tag in ("h1", "h2", "h3"):
            self._h_tag = tag
            self._h_text = []
        elif tag == "table":
            self.in_table = True
            self.current_table = []
        elif tag == "tr":
            self.in_row = True
            self.current_row = []
        elif tag in ("td", "th"):
            self.in_cell = True
            self.current_cell_text = []
        elif tag == "br":
            if self.in_cell:
                self.current_cell_text.append("\n")

    # ---- 标签结束 ----
    def handle_endtag(self, tag):
        if tag in ("h1", "h2", "h3"):
            title = "".join(self._h_text).strip()
            # 忽略封面/目录的 h3（图片或无文字的空标题）
            if title:
                self.current_section = title
            self._h_tag = None
            self._h_text = []
        elif tag == "table":
            self.in_table = False
            if self.current_section and self.current_table:
                self.sections.setdefault(self.current_section, []).append(self.current_table)
            self.current_table = []
        elif tag == "tr":
            self.in_row = False
            if self.current_row:
                self.current_table.append(self.current_row)
            self.current_row = []
        elif tag in ("td", "th"):
            self.in_cell = False
            text = "".join(self.current_cell_text).strip()
            self.current_row.append(text)
            self.current_cell_text = []

    # ---- 文本 ----
    def handle_data(self, data):
        if self._h_tag:
            self._h_text.append(data)
        elif self.in_cell:
            self.current_cell_text.append(data)


# ---------- 数据提取 ----------

def extract_industry_info(section_tables: list) -> dict:
    """
    从 1.2 工商信息 表格提取 {字段: 值}。
    表格行例如：
      ['企业名称', '广州奇圣电子科技股份有限公司']          (2 列, 后者 colspan)
      ['曾用名', '广州奇圣电子科技有限公司', '企业规模', 'XS（微型）']  (4 列)
    """
    info = {}
    for table in section_tables or []:
        for row in table:
            # 跳过表头行/空行
            if not row:
                continue
            if len(row) >= 4:
                for j in range(0, len(row) - 1, 2):
                    key = row[j].strip()
                    val = row[j + 1].strip()
                    if key and val:
                        info[key] = val
            elif len(row) == 2:
                key, val = row[0].strip(), row[1].strip()
                if key and val:
                    info[key] = val
    return info


def extract_table_rows(section_tables: list) -> list:
    """
    通用表格提取（股东信息 / 主要人员 / 变更记录）。
    返回 [{表头: 单元格值}, ...] 行字典列表。
    """
    rows_data = []
    for table in section_tables or []:
        if not table:
            continue
        header = table[0]
        for row in table[1:]:
            # 跳过全空行
            if not row or all(not c for c in row):
                continue
            item = {
                header[j]: row[j] if j < len(row) else ""
                for j in range(len(header))
            }
            rows_data.append(item)
    return rows_data


# ---------- 主流程 ----------

def parse_report(docx_path: str) -> dict:
    """解析报告文档，返回结构化信息"""
    if not os.path.isfile(docx_path):
        logger.error("文件不存在: %s", docx_path)
        raise FileNotFoundError(f"文件不存在: {docx_path}")

    logger.info("开始解析文档: %s", os.path.basename(docx_path))

    html_text = pypandoc.convert_file(docx_path, "html", format="docx")
    logger.debug("pandoc 转换 HTML 完成，长度 %d", len(html_text))

    parser = ReportHTMLParser()
    parser.feed(html_text)
    parser.close()
    logger.info("识别到 %d 个 HTML 章节标题", len(parser.sections))

    data = {"source_file": os.path.basename(docx_path)}

    # 按标题定位各章节包含的表格
    def section_tables(title_part):
        for sec_title, tables in parser.sections.items():
            if title_part in sec_title:
                logger.debug("章节 '%s' 匹配 '%s'，含 %d 个表格",
                             sec_title, title_part, len(tables))
                return tables
        logger.warning("未找到章节: %s", title_part)
        return None

    # 1.2 工商信息
    industry = extract_industry_info(section_tables("工商信息"))
    data["公司名称"] = industry.get("企业名称", "")
    data["法定代表人"] = industry.get("法定代表人", "")
    data["投资人"] = industry.get("投资人", "")
    data["经营者"] = industry.get("经营者", "")
    data["工商信息"] = industry
    logger.info("工商信息提取完成，共 %d 个字段（公司: %s，法人: %s，投资人: %s，经营者: %s）",
                len(industry), data["公司名称"], data["法定代表人"] or "-",
                data["投资人"] or "-", data["经营者"] or "-")

    # 1.3 股东信息
    shareholders = extract_table_rows(section_tables("股东信息"))
    data["股东"] = shareholders
    logger.info("股东信息提取完成，共 %d 条", len(shareholders))
    for sh in shareholders:
        logger.debug("  股东: %s 持股 %s 认缴 %s 万元",
                     sh.get("发起人名称"), sh.get("持股比例"),
                     sh.get("认缴出资额(万元)"))

    # 1.4 主要人员
    members = extract_table_rows(section_tables("主要人员"))
    data["主要人员"] = members
    logger.info("主要人员提取完成，共 %d 条", len(members))

    return data


def build_person_rows(data: dict) -> list:
    """
    以人名为索引汇总：
    法定代表人 / 股东 / 主要人员 -> 每个人一行。
    """
    company_name = data.get("公司名称", "")
    legal_person = data.get("法定代表人", "")

    persons = OrderedDict()

    def get_person(name: str) -> dict:
        p = persons.setdefault(name, {
            "姓名": name,
            "公司名称": company_name,
            "身份": [],
            "职务": [],
            "持股比例": [],
            "认缴出资额(万元)": "",
            "认缴出资日期": "",
            "首次持股日期": "",
        })
        return p

    # 法定代表人
    if legal_person and legal_person != "-":
        get_person(legal_person)["身份"].append("法定代表人")

    # 投资人（个人独资企业等）
    investor = data.get("投资人", "")
    if investor and investor != "-":
        get_person(investor)["身份"].append("投资人")

    # 经营者（个体工商户等）
    operator = data.get("经营者", "")
    if operator and operator != "-":
        get_person(operator)["身份"].append("经营者")

    # 股东
    for sh in data.get("股东", []):
        name = (sh.get("发起人名称") or sh.get("股东名称") or "").strip()
        if not name or name == "-":
            continue
        p = get_person(name)
        p["身份"].append("股东")
        ratio = (sh.get("持股比例") or "").strip()
        if ratio and ratio != "-":
            p["持股比例"].append(ratio)
        if sh.get("认缴出资额(万元)"):
            p["认缴出资额(万元)"] = sh["认缴出资额(万元)"]
        if sh.get("认缴出资日期"):
            p["认缴出资日期"] = sh["认缴出资日期"]
        if sh.get("首次持股日期"):
            p["首次持股日期"] = sh["首次持股日期"]

    # 主要人员
    for m in data.get("主要人员", []):
        name = (m.get("姓名") or "").strip()
        if not name or name == "-":
            continue
        p = get_person(name)
        p["身份"].append("主要成员")
        duty = (m.get("职务") or "").strip()
        if duty and duty != "-":
            p["职务"].append(duty)
        ratio = (m.get("持股比例") or "").strip()
        if ratio and ratio != "-":
            p["持股比例"].append(ratio)

    rows = []
    for name, p in persons.items():
        # 去重保序
        seen = set()
        roles = [r for r in p["身份"] if not (r in seen or seen.add(r))]
        duties = [d for d in p["职务"] if not (d in seen or seen.add(d))]
        ratios = [r for r in p["持股比例"] if not (r in seen or seen.add(r))]
        rows.append({
            "姓名": name,
            "公司名称": p["公司名称"],
            "身份": "、".join(roles),
            "职务": "、".join(duties),
            "持股比例": "、".join(ratios),
            "认缴出资额(万元)": p["认缴出资额(万元)"],
            "认缴出资日期": p["认缴出资日期"],
            "首次持股日期": p["首次持股日期"],
        })
        logger.debug("汇总人员: %s | 身份: %s | 职务: %s | 持股: %s",
                     name, "、".join(roles), "、".join(duties), "、".join(ratios))
    logger.info("人员汇总完成，共 %d 人", len(rows))
    return rows


def merge_all_companies(all_reports: list) -> list:
    """
    合并多份报告的人员信息：
    以人名为索引，同一人出现在多家公司时，
    各信息列用换行分隔对应各公司。
    """
    # 姓名 -> {列: [值, 值, ...]}（每个值对应一家公司）
    person_map = OrderedDict()
    # 身份统计：身份 -> 人数
    role_stats = OrderedDict()
    # 公司统计
    company_count = 0
    company_names = set()
    # 投资人统计
    investor_count = 0
    investor_names = set()
    # 经营者统计
    operator_count = 0
    operator_names = set()
    # 法定代表人统计
    legal_count = 0
    legal_names = set()
    # 股东统计
    shareholder_count = 0
    shareholder_names = set()
    # 主要成员统计
    member_count = 0
    member_names = set()
    # 文件统计
    total_files = len(all_reports)
    fail_count = 0
    # 各公司解析状态
    company_status = []

    for report in all_reports:
        company = report.get("公司名称", "") or report.get("source_file", "")
        legal = report.get("法定代表人", "")
        investor = report.get("投资人", "")
        operator = report.get("经营者", "")
        shareholders = report.get("股东", [])
        members = report.get("主要人员", [])

        # 统计公司
        if company:
            company_count += 1
            company_names.add(company)

        # 统计法定代表人
        if legal and legal != "-":
            legal_count += 1
            legal_names.add(legal)

        # 统计投资人
        if investor and investor != "-":
            investor_count += 1
            investor_names.add(investor)

        # 统计经营者
        if operator and operator != "-":
            operator_count += 1
            operator_names.add(operator)

        # 统计股东
        for sh in shareholders:
            name = (sh.get("发起人名称") or sh.get("股东名称") or "").strip()
            if name and name != "-":
                shareholder_count += 1
                shareholder_names.add(name)

        # 统计主要成员
        for m in members:
            name = (m.get("姓名") or "").strip()
            if name and name != "-":
                member_count += 1
                member_names.add(name)

        # 收集该公司下所有人，整理该人在该公司的信息
        company_persons = OrderedDict()

        def add_person(name: str):
            if not name or name == "-":
                return None
            p = company_persons.setdefault(name, {
                "身份": [], "职务": [], "持股比例": [],
                "认缴出资额(万元)": "", "认缴出资日期": "", "首次持股日期": "",
            })
            return p

        # 法定代表人
        if legal and legal != "-":
            p = add_person(legal)
            if p:
                p["身份"].append("法定代表人")

        # 投资人（个人独资企业等）
        if investor and investor != "-":
            p = add_person(investor)
            if p:
                p["身份"].append("投资人")

        # 经营者（个体工商户等）
        if operator and operator != "-":
            p = add_person(operator)
            if p:
                p["身份"].append("经营者")

        # 股东
        for sh in shareholders:
            name = (sh.get("发起人名称") or sh.get("股东名称") or "").strip()
            p = add_person(name)
            if p is None:
                continue
            p["身份"].append("股东")
            ratio = (sh.get("持股比例") or "").strip()
            if ratio and ratio != "-":
                p["持股比例"].append(ratio)
            for k in ("认缴出资额(万元)", "认缴出资日期", "首次持股日期"):
                v = (sh.get(k) or "").strip()
                if v and v != "-" and not p[k]:
                    p[k] = v

        # 主要人员
        for m in members:
            name = (m.get("姓名") or "").strip()
            p = add_person(name)
            if p is None:
                continue
            p["身份"].append("主要成员")
            duty = (m.get("职务") or "").strip()
            if duty and duty != "-":
                p["职务"].append(duty)
            ratio = (m.get("持股比例") or "").strip()
            if ratio and ratio != "-":
                p["持股比例"].append(ratio)

        # 写入全局索引
        for name, p in company_persons.items():
            g = person_map.setdefault(name, {
                "姓名": name,
                "公司名称": [], "身份": [], "职务": [],
                "持股比例": [], "认缴出资额(万元)": [],
                "认缴出资日期": [], "首次持股日期": [],
            })
            g["公司名称"].append(company)
            g["身份"].append("、".join(dict.fromkeys(p["身份"])))
            g["职务"].append("、".join(dict.fromkeys(p["职务"])))
            g["持股比例"].append("、".join(dict.fromkeys(p["持股比例"])))
            g["认缴出资额(万元)"].append(p["认缴出资额(万元)"])
            g["认缴出资日期"].append(p["认缴出资日期"])
            g["首次持股日期"].append(p["首次持股日期"])

    # 汇总输出，多值用换行连接
    rows = []
    for name, g in person_map.items():
        rows.append({
            "姓名": name,
            "公司名称": "\n".join(g["公司名称"]),
            "身份": "\n".join(g["身份"]),
            "职务": "\n".join(g["职务"]),
            "持股比例": "\n".join(g["持股比例"]),
            "认缴出资额(万元)": "\n".join(g["认缴出资额(万元)"]),
            "认缴出资日期": "\n".join(g["认缴出资日期"]),
            "首次持股日期": "\n".join(g["首次持股日期"]),
        })
        logger.debug("合并: %s | %d 家公司", name, len(g["公司名称"]))
    logger.info("跨公司人员合并完成，共 %d 人", len(rows))

    # 构建统计信息
    stats = {
        "total_files": total_files,
        "company_count": company_count,
        "company_names": sorted(company_names),
        "legal_count": legal_count,
        "legal_names": sorted(legal_names),
        "investor_count": investor_count,
        "investor_names": sorted(investor_names),
        "operator_count": operator_count,
        "operator_names": sorted(operator_names),
        "shareholder_count": shareholder_count,
        "shareholder_names": sorted(shareholder_names),
        "member_count": member_count,
        "member_names": sorted(member_names),
        "person_count": len(rows),
        "fail_count": fail_count,
        "company_details": [],
    }

    # 每个公司的详细统计
    for report in all_reports:
        company = report.get("公司名称", "") or report.get("source_file", "")
        legal = report.get("法定代表人", "")
        investor = report.get("投资人", "")
        operator = report.get("经营者", "")
        shareholders = report.get("股东", [])
        members = report.get("主要人员", [])

        # 统计该公司下的股东和主要成员
        sh_names = set()
        for sh in shareholders:
            name = (sh.get("发起人名称") or sh.get("股东名称") or "").strip()
            if name and name != "-":
                sh_names.add(name)

        m_names = set()
        for m in members:
            name = (m.get("姓名") or "").strip()
            if name and name != "-":
                m_names.add(name)

        # 该公司所有人员（去重）
        all_persons = set()
        if legal and legal != "-":
            all_persons.add(legal)
        if investor and investor != "-":
            all_persons.add(investor)
        if operator and operator != "-":
            all_persons.add(operator)
        all_persons.update(sh_names)
        all_persons.update(m_names)

        stats["company_details"].append({
            "company": company,
            "legal": legal if legal and legal != "-" else "",
            "investor": investor if investor and investor != "-" else "",
            "operator": operator if operator and operator != "-" else "",
            "shareholders": sorted(sh_names),
            "members": sorted(m_names),
            "person_count": len(all_persons),
        })
    logger.info("统计: %d 家公司 / %d 位人员 / %d 位法人 / %d 位投资人 / %d 位经营者 / %d 位股东 / %d 位主要成员",
                company_count, len(rows), legal_count, investor_count,
                operator_count, shareholder_count, member_count)

    # 将统计信息附加到返回结果（通过全局变量或返回元组）
    # 这里使用模块级变量以便 GUI 获取
    global _last_stats
    _last_stats = stats

    return rows


# 模块级变量：保存最近一次合并的统计信息
_last_stats = None


def get_last_stats() -> dict:
    """获取最近一次合并的统计信息"""
    return _last_stats or {}


def save_xlsx(rows: list, out_path: str, stats: dict = None):
    """将人员行写入 xlsx，第2个 sheet 为各公司统计"""
    headers = ["姓名", "公司名称", "身份", "职务", "持股比例",
               "认缴出资额(万元)", "认缴出资日期", "首次持股日期"]

    wb = Workbook()
    ws = wb.active
    ws.title = "企业人员信息"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 数据行
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(h, ""))
            cell.alignment = Alignment(vertical="center")

    # 列宽
    col_widths = [12, 38, 22, 30, 16, 18, 16, 16]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 冻结首行
    ws.freeze_panes = "A2"

    # 自动筛选
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows) + 1}"

    # ===== 第2个 sheet：各公司统计 =====
    if stats:
        ws2 = wb.create_sheet("各公司统计")
        stats_headers = ["序号", "公司名称", "法定代表人", "投资人", "经营者",
                         "股东人数", "主要成员人数", "人员总数"]
        for col_idx, h in enumerate(stats_headers, start=1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        company_details = stats.get("company_details", [])
        for row_idx, cd in enumerate(company_details, start=2):
            ws2.cell(row=row_idx, column=1, value=row_idx - 1)
            ws2.cell(row=row_idx, column=2, value=cd.get("company", ""))
            ws2.cell(row=row_idx, column=3, value=cd.get("legal", "") or "-")
            ws2.cell(row=row_idx, column=4, value=cd.get("investor", "") or "-")
            ws2.cell(row=row_idx, column=5, value=cd.get("operator", "") or "-")
            ws2.cell(row=row_idx, column=6, value=len(cd.get("shareholders", [])))
            ws2.cell(row=row_idx, column=7, value=len(cd.get("members", [])))
            ws2.cell(row=row_idx, column=8, value=cd.get("person_count", 0))
            for col_idx in range(1, len(stats_headers) + 1):
                ws2.cell(row=row_idx, column=col_idx).alignment = Alignment(vertical="center")

        # 列宽
        stats_col_widths = [6, 40, 14, 14, 14, 10, 14, 10]
        for i, w in enumerate(stats_col_widths, start=1):
            ws2.column_dimensions[get_column_letter(i)].width = w

        # 冻结首行
        ws2.freeze_panes = "A2"

        # 自动筛选
        ws2.auto_filter.ref = f"A1:{get_column_letter(len(stats_headers))}{len(company_details) + 1}"

    wb.save(out_path)
    logger.info("xlsx 保存完成: %s（%d 行 × %d 列 + 统计 sheet）",
                out_path, len(rows), len(headers))


# ---------- 入口 ----------

def main():
    parser = argparse.ArgumentParser(
        description="解析企业信用报告 docx -> xlsx（人名为索引）"
    )
    parser.add_argument(
        "input",
        nargs="?",
        default=r"D:\Download\word",
        help="Word 文档路径 或 目录路径（默认 D:\\Download\\word）",
    )
    parser.add_argument(
        "-o", "--output",
        default="company_report.xlsx",
        help="输出 xlsx 文件路径（默认 company_report.xlsx）",
    )
    parser.add_argument(
        "--json", action="store_true",
        help="同时输出中间 JSON（便于检查）",
    )
    parser.add_argument(
        "--log", default="company_doc.log",
        help="日志文件路径（默认 company_doc.log）",
    )
    parser.add_argument(
        "-v", "--verbose", action="store_true",
        help="在控制台输出详细信息（默认只显示警告与错误）",
    )
    args = parser.parse_args()

    # 配置日志
    setup_logging(args.log, verbose=args.verbose)
    logger.info("========== 企业信用报告解析开始 ==========")
    logger.info("输入: %s", args.input)
    logger.info("输出文件: %s", args.output)
    logger.info("日志文件: %s", args.log)

    # 判断输入是文件还是目录
    input_path = Path(args.input)
    if input_path.is_dir():
        docx_files = sorted(input_path.glob("*.docx"))
        logger.info("目录解析模式，共 %d 个 docx 文件", len(docx_files))
    elif input_path.is_file():
        docx_files = [input_path]
        logger.info("单文件解析模式")
    else:
        logger.error("输入路径不存在: %s", args.input)
        sys.exit(f"错误: 输入路径不存在: {args.input}")

    if not docx_files:
        logger.warning("未找到任何 docx 文件")
        sys.exit("错误: 未找到任何 docx 文件")

    # 逐个解析
    all_reports = []
    fail_count = 0
    for f in docx_files:
        try:
            data = parse_report(str(f))
            # 保存单文件详细字段（公司名、法人等），供 JSON 输出
            all_reports.append(data)
        except Exception as e:
            fail_count += 1
            logger.error("解析失败 %s: %s", f.name, e)

    if not all_reports:
        logger.error("所有文件均解析失败（共 %d 个）", fail_count)
        sys.exit(f"错误: 所有 {docx_files.__len__()} 个文件均解析失败")

    logger.info("解析完成：成功 %d 个，失败 %d 个", len(all_reports), fail_count)

    # 合并所有公司的人员信息（人名为索引）
    rows = merge_all_companies(all_reports)

    # 获取统计信息
    stats = get_last_stats()

    # 保存 xlsx（含第2个 sheet 统计）
    try:
        save_xlsx(rows, args.output, stats=stats)
    except Exception as e:
        logger.exception("xlsx 保存失败")
        sys.exit(f"错误: xlsx 保存失败: {e}")

    out_path = Path(args.output)
    print(f"解析完成，已保存到: {out_path.resolve()}")
    print(f"共解析 {len(all_reports)} 家公司 / {len(rows)} 位人员（失败 {fail_count}）")
    logger.info("解析完成，共 %d 家公司 / %d 位人员，输出到 %s",
                len(all_reports), len(rows), out_path.resolve())

    if args.json:
        json_path = out_path.with_suffix(".json")
        json_path.write_text(
            json.dumps(all_reports, ensure_ascii=False, indent=2, default=str),
            encoding="utf-8",
        )
        print(f"中间 JSON: {json_path.resolve()}")
        logger.info("中间 JSON 已保存: %s", json_path.resolve())

    logger.info("========== 企业信用报告解析结束 ==========")


if __name__ == "__main__":
    main()
